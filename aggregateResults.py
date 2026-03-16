import argparse
import csv
import os
import numpy as np
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def try_float(val) -> float:
    try:
        return float(str(val).strip())
    except Exception:
        return float("nan")

def try_int(val) -> int:
    try:
        return int(float(str(val).strip()))
    except Exception:
        return 0

def is_percentile_col(col_name: str) -> bool:
    n = str(col_name).strip()
    if "/s" in n:
        return False
    if not n.endswith("%"):
        return False
    try:
        float(n[:-1])
        return True
    except ValueError:
        return False

def convert_ms_to_s(rows: List[Dict[str, Any]], fieldnames: List[str]):
    out_fieldnames = []
    convert_map = {}
    for c in fieldnames:
        if is_percentile_col(c):
            new_name = f"{c} (s)"
            out_fieldnames.append(new_name)
            convert_map[c] = new_name
        else:
            out_fieldnames.append(c)

    new_rows = []
    for r in rows:
        nr = {}
        for orig in fieldnames:
            if orig in convert_map:
                fv = try_float(r.get(orig, ""))
                nr[convert_map[orig]] = f"{(fv / 1000.0):.2f}" if np.isfinite(fv) else ""
            else:
                nr[orig] = r.get(orig, "")
        new_rows.append(nr)
    return new_rows, out_fieldnames

def aggregate_events(rows: List[Dict[str, Any]], out_dir: str):
    percentile_cols = [c for c in rows[0].keys() if c.endswith(" (s)") and is_percentile_col(c[:-4]) and (c.startswith("50%") or c.startswith("99%"))]
    groups = {}
    for r in rows:
        if (r.get("Type") or "").strip().lower() != "event":
            continue
        name = (r.get("Name") or "").strip()
        if name.lower() == "aggregated" or name in ["99-InterIterationDelay", "NaturalPause"]:
            continue
        req = try_int(r.get("Request Count"))
        fail = try_int(r.get("Failure Count"))
        succ = max(req - fail, 0)
        g = groups.setdefault(name, {"req": 0, "fail": 0, "pct_vals": {pc: [] for pc in percentile_cols}})
        g["req"] += req
        g["fail"] += fail
        for pc in percentile_cols:
            pv = try_float(r.get(pc))
            if np.isfinite(pv) and succ > 0:
                g["pct_vals"][pc].append((pv, succ))

    wb = Workbook()
    ws = wb.active
    fieldnames = ["Name", "Total Request Count", "Total Failure Count", "Failure Rate"] + percentile_cols + ["Benchmark Status"]
    ws.append(fieldnames)

    preferred_order = ["LaunchUrl", "TypeCredentials", "Login", "SignOut"]
    sorted_names = preferred_order + sorted([n for n in groups.keys() if n not in preferred_order])

    # --- track totals for summary row ---
    total_req_sum = 0
    total_fail_sum = 0
    failure_rates = []
    pct_summary_vals = {pc: [] for pc in percentile_cols}

    for name in sorted_names:
        g = groups.get(name)
        if g is None:
            continue
        req, fail = g["req"], g["fail"]
        fail_rate = (fail / req) if req > 0 else 0
        row = [name, req, fail, f"{fail_rate:.6f}"]

        total_req_sum += req
        total_fail_sum += fail
        if req > 0:
            failure_rates.append(fail_rate)

        for pc in percentile_cols:
            vals_wts = g["pct_vals"][pc]
            if vals_wts:
                vals = np.array([vw[0] for vw in vals_wts])
                wts = np.array([vw[1] for vw in vals_wts])
                avg_val = np.average(vals, weights=wts)
                row.append(f"{avg_val:.2f}")
                pct_summary_vals[pc].extend([(v, w) for v, w in vals_wts])
            else:
                row.append("")

        p50 = try_float(row[fieldnames.index("50% (s)")]) if "50% (s)" in fieldnames else float("nan")
        p99 = try_float(row[fieldnames.index("99% (s)")]) if "99% (s)" in fieldnames else float("nan")
        if np.isfinite(p50) and np.isfinite(p99):
            if p50 <= 1.0 and p99 <= 2.0:
                status = "PASSED"
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                status = "FAILED"
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else:
            status = "NO DATA"
            fill = None

        row.append(status)
        ws.append(row)
        if fill:
            status_col = fieldnames.index("Benchmark Status") + 1
            ws.cell(row=ws.max_row, column=status_col).fill = fill

    # --- add summary row at bottom ---
    avg_failure_rate = np.mean(failure_rates) if failure_rates else 0.0
    fail_pct = (total_fail_sum / total_req_sum * 100.0) if total_req_sum > 0 else 0.0
    summary_row = [
        "TOTALS",
        total_req_sum,
        f"{total_fail_sum} ({fail_pct:.2f}%)",
        f"{avg_failure_rate:.6f}",
    ]

    # add weighted averages for percentile columns
    for pc in percentile_cols:
        vals_wts = pct_summary_vals[pc]
        if vals_wts:
            vals = np.array([vw[0] for vw in vals_wts])
            wts = np.array([vw[1] for vw in vals_wts])
            summary_row.append(f"{np.average(vals, weights=wts):.2f}")
        else:
            summary_row.append("")

    # add blank for Benchmark Status
    summary_row.append("")

    ws.append(summary_row)

    # Style summary row
    for col in range(1, len(summary_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
    ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    os.makedirs(out_dir, exist_ok=True)
    wb.save(os.path.join(out_dir, "events_aggregated.xlsx"))
    print(f"[OK] Aggregated events written to: {os.path.join(out_dir, 'events_aggregated.xlsx')}")

def main():
    parser = argparse.ArgumentParser(description="Aggregate Locust summary CSV with benchmark check.")
    parser.add_argument("--in", dest="input_file", required=True, help="Path to Locust summary CSV.")
    parser.add_argument("--out-dir", dest="out_dir", default=".", help="Output directory.")
    args = parser.parse_args()

    with open(args.input_file, "r", encoding="utf-8-sig", newline="") as f:
        rdr = csv.DictReader(f)
        rows = [r for r in rdr]
        fieldnames = rdr.fieldnames or []

    rows_s, _ = convert_ms_to_s(rows, fieldnames)
    aggregate_events(rows_s, args.out_dir)

if __name__ == "__main__":
    main()
